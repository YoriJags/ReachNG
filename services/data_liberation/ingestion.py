"""
File ingestion for Data Liberation.
Accepts PDF, Excel, CSV, and plain text.
Extracts text, splits into chunks, saves to MongoDB.
"""
import io
import csv
import structlog
from .store import save_chunks

log = structlog.get_logger()

# Max characters per chunk — keeps each chunk well within Claude's context
CHUNK_SIZE = 1500
CHUNK_OVERLAP = 200


async def ingest_file(
    client_name: str,
    filename: str,
    content: bytes,
) -> dict:
    """
    Ingest a file for a client. Detects type, extracts text, chunks, saves.
    Returns {source_file, file_type, chunks_saved, status}
    """
    ext = filename.lower().rsplit(".", 1)[-1] if "." in filename else "txt"

    try:
        if ext == "pdf":
            chunks = _extract_pdf(content, filename)
            file_type = "pdf"
        elif ext in ("xlsx", "xls"):
            chunks = _extract_excel(content, filename)
            file_type = "excel"
        elif ext == "csv":
            chunks = _extract_csv(content, filename)
            file_type = "csv"
        else:
            # Plain text, WhatsApp export, etc.
            chunks = _extract_text(content.decode("utf-8", errors="replace"), filename)
            file_type = "text"

        count = save_chunks(client_name, filename, file_type, chunks)
        log.info("ingest_complete", client=client_name, file=filename, chunks=count)
        return {"source_file": filename, "file_type": file_type, "chunks_saved": count, "status": "ok"}

    except Exception as e:
        log.error("ingest_failed", client=client_name, file=filename, error=str(e))
        return {"source_file": filename, "file_type": ext, "chunks_saved": 0, "status": "error", "error": str(e)}


# ─── Extractors ───────────────────────────────────────────────────────────────

def _extract_pdf(content: bytes, filename: str) -> list[dict]:
    try:
        import pdfplumber
        chunks = []
        with pdfplumber.open(io.BytesIO(content)) as pdf:
            for page_num, page in enumerate(pdf.pages, 1):
                text = page.extract_text() or ""
                if text.strip():
                    for chunk in _split_text(text):
                        chunks.append({
                            "content": chunk,
                            "metadata": {"page": page_num, "source": filename},
                        })
        return chunks
    except ImportError:
        # Fallback: read raw bytes as text (catches text-based PDFs partially)
        log.warning("pdfplumber_not_installed_falling_back")
        text = content.decode("latin-1", errors="replace")
        return [{"content": c, "metadata": {"source": filename}} for c in _split_text(text)]


def _extract_excel(content: bytes, filename: str) -> list[dict]:
    try:
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        chunks = []
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            rows = []
            for row in ws.iter_rows(values_only=True):
                row_text = " | ".join(str(cell) for cell in row if cell is not None)
                if row_text.strip():
                    rows.append(row_text)
            # Group rows into chunks
            for i in range(0, len(rows), 30):
                batch = rows[i:i + 30]
                chunks.append({
                    "content": f"Sheet: {sheet}\n" + "\n".join(batch),
                    "metadata": {"sheet": sheet, "rows": f"{i+1}-{i+len(batch)}", "source": filename},
                })
        return chunks
    except ImportError:
        log.warning("openpyxl_not_installed")
        return [{"content": f"[Excel file: {filename} — openpyxl not installed]", "metadata": {}}]


def _extract_csv(content: bytes, filename: str) -> list[dict]:
    text = content.decode("utf-8", errors="replace")
    reader = csv.reader(io.StringIO(text))
    rows = [" | ".join(row) for row in reader if any(cell.strip() for cell in row)]
    chunks = []
    for i in range(0, len(rows), 30):
        batch = rows[i:i + 30]
        chunks.append({
            "content": "\n".join(batch),
            "metadata": {"rows": f"{i+1}-{i+len(batch)}", "source": filename},
        })
    return chunks


def _extract_text(text: str, filename: str) -> list[dict]:
    return [
        {"content": chunk, "metadata": {"source": filename}}
        for chunk in _split_text(text)
    ]


def _split_text(text: str) -> list[str]:
    """Split text into overlapping chunks."""
    chunks = []
    start = 0
    while start < len(text):
        end = start + CHUNK_SIZE
        chunks.append(text[start:end].strip())
        start += CHUNK_SIZE - CHUNK_OVERLAP
    return [c for c in chunks if c]
